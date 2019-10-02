import openpyxl as xl
from openpyxl.chart import BarChart, Reference

##### below code is not dynamic. because I specified fine name.
# wb = xl.load_workbook("transactions.xlsx")
# sheet = wb["Sheet1"]
#
# ## 2 ways to pick column and row
# ##both of them are getting cell of column A / row 1 or (row, column)
# # cell = sheet['b1']
# # cell2 = sheet.cell(1, 2)
# # print(cell.value)
# # print(cell2.value)
#
# ###access to current price on the each row and calculate 10% disccounted price
#
# ##check the number of row
#
# for row in range(2, sheet.max_row + 1):
#     cur_cell = sheet.cell(row, 3).value
#     sheet.cell(row, 5).value = cur_cell * 0.9
#
# values = Reference(sheet,
#           min_row = 2,
#           max_row=sheet.max_row,
#           min_col=4,
#           max_col=4)
#
# chart = BarChart()
# chart.add_data(values)
# sheet.add_chart(chart, 'e2')
#
# ## instead of overwritting the file, I created v2
# wb.save('transactions2.xlsx')

###below is the dynamic version

def process_workbook(workbook_name, sheet_num):
    wb = xl.load_workbook(workbook_name)
    sheet = wb[sheet_num]

    for row in range(2, sheet.max_row + 1):
        cur_cell = sheet.cell(row, 3).value
        sheet.cell(row, 5).value = cur_cell * 0.9

    values = Reference(sheet,
              min_row = 2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'g2')

    wb.save(workbook_name)

process_workbook("transactions.xlsx", "Sheet1")